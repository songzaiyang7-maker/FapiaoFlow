"""Excel 导出。

输出文件结构：
- Sheet "发票明细"：每张发票一行，列与 UI 表格一致
- Sheet "汇总"：各类小计 + 总计 + 未分类（如有）

文件名格式：fapiao_summary_YYYYMMDD_HHMMSS.xlsx
"""

from __future__ import annotations

import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.core.categories import CategoryDef
from src.core.totals import compute_totals
from src.core.types import Invoice

DETAIL_HEADERS = ["文件名", "类目", "金额", "发票号", "销售方", "日期", "状态/备注"]


def export_to_excel(
    invoices: list[Invoice],
    categories: list[CategoryDef],
    output_path: str | Path | None = None,
) -> Path:
    """导出到 .xlsx。返回文件路径。

    output_path 为 None 时生成默认文件名到当前目录。
    """
    if output_path is None:
        now = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = Path.cwd() / f"fapiao_summary_{now}.xlsx"
    output_path = Path(output_path)

    wb = Workbook()

    # Sheet 1: 发票明细
    ws = wb.active
    ws.title = "发票明细"
    ws.append(DETAIL_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4A6FFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for inv in invoices:
        if inv.error:
            status = "❌ " + inv.error
        elif inv.user_overridden:
            status = "用户已改"
        elif inv.confidence < 0.7:
            status = "⚠ 待复核"
        else:
            status = "已分类"
        ws.append([
            inv.file_name,
            inv.category or "",
            inv.amount if inv.amount is not None else "",
            inv.invoice_no or "",
            inv.seller_name or "",
            inv.issue_date or "",
            status,
        ])

    # 列宽：用 get_column_letter 避免 chr(64+i) 超过 26 列出错
    widths = [25, 10, 12, 22, 30, 12, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Sheet 2: 汇总（类目从传入参数动态生成）
    ws2 = wb.create_sheet("汇总")
    totals = compute_totals(invoices, categories)
    ws2.append(["类目", "金额（元）"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)

    sorted_cats = sorted(categories, key=lambda c: (c.priority, c.name))
    summary_row = 2
    for cat in sorted_cats:
        ws2.append([cat.name, totals[cat.name]])
        summary_row += 1
    # 未分类（如果有发票指向已删除类目）
    if totals.get("未分类", 0.0) != 0.0:
        ws2.append(["未分类", totals["未分类"]])
        summary_row += 1
    # 总计行
    ws2.append(["总计", totals["总计"]])
    ws2.cell(row=summary_row + 1, column=1).font = Font(bold=True)
    ws2.cell(row=summary_row + 1, column=2).font = Font(bold=True, color="FF0000")
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 14

    wb.save(str(output_path))
    return output_path
